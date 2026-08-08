from __future__ import annotations

import argparse
import bisect
import csv
import os
from dataclasses import dataclass
from itertools import islice
from pathlib import Path
from typing import Any

from rich.text import Text
from textual.app import App, ComposeResult
from textual.binding import Binding
from textual.containers import Horizontal, Vertical
from textual.events import Key, MouseDown, MouseMove, MouseUp
from textual.widgets import DataTable, Input, Static


SUPPORTED_EXTENSIONS = {
    ".csv",
    ".data",
    ".txt",
    ".tsv",
    ".xlsx",
    ".xls",
    ".parquet",
}

HEAVY_DIRS = {
    ".git",
    ".ruff_cache",
    ".pytest_cache",
    ".mypy_cache",
    ".venv",
    "venv",
    "__pycache__",
    "node_modules",
}

FOLDER_CLOSED_ICON = ""
FOLDER_OPEN_ICON = ""
FILE_DEFAULT_ICON = ""
FILE_TABLE_ICON = ""
FILE_PYTHON_ICON = ""
FILE_MARKDOWN_ICON = ""
FILE_CONFIG_ICON = ""
FILE_GIT_ICON = ""
MIN_TREE_PANEL_WIDTH = 20
MIN_DATA_PANEL_WIDTH = 40


@dataclass(frozen=True)
class FileRow:
    path: Path
    ancestors_last: tuple[bool, ...]


class PaneDivider(Static):
    """A draggable divider between the file tree and data view."""

    def on_mouse_down(self, event: MouseDown) -> None:
        if event.button != 1:
            return

        self.app.begin_tree_resize(event.screen_x)
        self.capture_mouse()
        self.add_class("dragging")
        event.stop()

    def on_mouse_move(self, event: MouseMove) -> None:
        if self.app.mouse_captured is not self:
            return

        self.app.resize_tree_from_pointer(event.screen_x)
        event.stop()

    def on_mouse_up(self, event: MouseUp) -> None:
        if self.app.mouse_captured is not self:
            return

        self.release_mouse()
        self.remove_class("dragging")
        event.stop()


class BaseWindowReader:
    columns: list[str]
    total_rows: int | None

    def read_window(self, offset: int, limit: int) -> list[list[Any]]:
        raise NotImplementedError


class DelimitedWindowReader(BaseWindowReader):
    def __init__(self, path: Path, index_step: int = 500) -> None:
        self.path = path
        self.index_step = max(index_step, 1)
        self.dialect = self._detect_dialect()
        self.columns: list[str] = []
        self.total_rows: int | None = None
        self._row_offsets: list[int] = []
        self._row_numbers: list[int] = []
        self._build_index()

    def _detect_dialect(self) -> csv.Dialect:
        with self.path.open("r", encoding="utf-8-sig", newline="") as file:
            sample = file.read(8192)

        if self.path.suffix.lower() == ".tsv":
            return csv.excel_tab

        try:
            return csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            return csv.excel

    def _build_index(self) -> None:
        with self.path.open("r", encoding="utf-8-sig", newline="") as file:
            header_line = file.readline()

            if not header_line:
                self.columns = []
                self.total_rows = 0
                return

            self.columns = next(csv.reader([header_line], self.dialect))

            row_number = 0

            while True:
                position = file.tell()
                line = file.readline()

                if not line:
                    break

                if row_number % self.index_step == 0:
                    self._row_numbers.append(row_number)
                    self._row_offsets.append(position)

                row_number += 1

            self.total_rows = row_number

    def read_window(self, offset: int, limit: int) -> list[list[Any]]:
        if self.total_rows is None or self.total_rows == 0:
            return []

        offset = max(0, min(offset, self.total_rows - 1))

        anchor_index = bisect.bisect_right(self._row_numbers, offset) - 1
        anchor_index = max(anchor_index, 0)

        anchor_row = self._row_numbers[anchor_index]
        anchor_position = self._row_offsets[anchor_index]
        rows_to_skip = offset - anchor_row

        with self.path.open("r", encoding="utf-8-sig", newline="") as file:
            file.seek(anchor_position)

            reader = csv.reader(file, self.dialect)

            for _ in range(rows_to_skip):
                next(reader, None)

            rows = list(islice(reader, limit))

        return [self._normalize_row(row) for row in rows]

    def _normalize_row(self, row: list[Any]) -> list[Any]:
        expected = len(self.columns)

        if len(row) < expected:
            return row + [""] * (expected - len(row))

        if len(row) > expected:
            return row[: expected - 1] + [",".join(row[expected - 1 :])]

        return row


class ExcelWindowReader(BaseWindowReader):
    def __init__(self, path: Path) -> None:
        from openpyxl import load_workbook

        self.workbook = load_workbook(path, read_only=True, data_only=True)
        self.sheet = self.workbook.active

        header_row = next(
            self.sheet.iter_rows(min_row=1, max_row=1, values_only=True),
            None,
        )

        self.columns = [
            str(value) if value is not None else f"column_{index + 1}"
            for index, value in enumerate(header_row or [])
        ]

        self.total_rows = max(self.sheet.max_row - 1, 0)

    def read_window(self, offset: int, limit: int) -> list[list[Any]]:
        if not self.total_rows:
            return []

        offset = max(0, min(offset, self.total_rows - 1))

        min_row = offset + 2
        max_row = min(offset + limit + 1, self.sheet.max_row)

        return [
            list(row)
            for row in self.sheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                values_only=True,
            )
        ]


class ParquetWindowReader(BaseWindowReader):
    def __init__(self, path: Path) -> None:
        import pandas as pd

        self.df = pd.read_parquet(path)
        self.columns = [str(col) for col in self.df.columns]
        self.total_rows = len(self.df)

    def read_window(self, offset: int, limit: int) -> list[list[Any]]:
        if self.total_rows == 0:
            return []

        offset = max(0, min(offset, self.total_rows - 1))

        return self.df.iloc[offset : offset + limit].values.tolist()


def create_reader(path: Path, window_size: int) -> BaseWindowReader:
    suffix = path.suffix.lower()

    if suffix in {".csv", ".data", ".txt", ".tsv"}:
        return DelimitedWindowReader(path, index_step=window_size)

    if suffix in {".xlsx", ".xls"}:
        return ExcelWindowReader(path)

    if suffix == ".parquet":
        return ParquetWindowReader(path)

    raise ValueError(f"Unsupported file extension: {suffix}")


def collect_supported_files(root: Path) -> list[Path]:
    files: list[Path] = []

    for current_root, dirs, filenames in os.walk(root):
        dirs[:] = [directory for directory in dirs if directory not in HEAVY_DIRS]
        current_path = Path(current_root)

        for filename in filenames:
            path = current_path / filename

            if path.suffix.lower() in SUPPORTED_EXTENSIONS:
                files.append(path.resolve())

    return sorted(files, key=lambda p: str(p).lower())


class TennisDataTUI(App):
    CSS = """
    Screen {
        background: #070b12;
        color: #cdd6f4;
        padding: 0;
        margin: 0;
    }

    #root {
        height: 100%;
        width: 100%;
        padding: 0;
        margin: 0;
        background: #070b12;
    }

    #main {
        height: 1fr;
        width: 100%;
        layout: horizontal;
        padding: 0;
        margin: 0;
        background: #070b12;
    }

    #tree_panel {
        width: 31%;
        min-width: 20;
        height: 100%;
        border: none;
        padding: 0;
        margin: 0;
        background: #0d111c;
    }

    #tree_header, #data_header {
        height: 1;
        padding: 0 1;
        margin: 0;
        background: #111827;
        color: #cdd6f4;
        text-style: bold;
    }

    #tree_header {
        color: #89b4fa;
    }

    #data_header {
        color: #a6e3a1;
    }

    #pane_divider {
        width: 1;
        min-width: 1;
        height: 100%;
        margin: 0;
        padding: 0;
        background: #313244;
    }

    #pane_divider:hover, #pane_divider.dragging {
        background: #89b4fa;
    }

    #data_panel {
        width: 1fr;
        min-width: 40;
        height: 100%;
        border: none;
        padding: 0;
        margin: 0;
        background: #0d111c;
    }

    #file_table {
        height: 1fr;
        width: 100%;
        background: #0d111c;
        color: #cdd6f4;
        scrollbar-size: 0 0;
    }

    #file_table > .datatable--header {
        height: 0;
        display: none;
    }

    #file_table > .datatable--cursor {
        background: #313244;
        color: #f9e2af;
    }

    #status {
        height: 4;
        padding: 0 0 1 0;
        margin: 0;
        color: #bac2de;
        border-bottom: solid #313244;
        background: #0d111c;
    }

    #data_table {
        height: 1fr;
        width: 100%;
        background: #0d111c;
        color: #cdd6f4;
        scrollbar-size: 0 0;
    }

    #data_table > .datatable--header {
        background: #313244;
        color: #f5e0dc;
        text-style: bold;
    }

    #data_table > .datatable--cursor {
        background: #45475a;
        color: #ffffff;
    }

    #command {
        height: 3;
        border: round #7f849c;
        background: #070b12;
        color: #cdd6f4;
        margin: 0;
    }

    #help_bar {
        height: 1;
        padding: 0 1;
        margin: 0;
        background: #070b12;
        color: #a6adc8;
    }
    """

    BINDINGS = [
        Binding("q", "quit", "Quit", priority=True),
        Binding("r", "reload", "Reload", priority=True),
        Binding("h", "toggle_left", "Toggle tree", priority=True),
        Binding("n", "next_window", "Next chunk", priority=True),
        Binding("p", "previous_window", "Previous chunk", priority=True),
        Binding("pagedown", "next_window", "Next chunk", priority=True),
        Binding("pageup", "previous_window", "Previous chunk", priority=True),
        Binding("g", "first_window", "First rows", priority=True),
        Binding("shift+g", "last_window", "Last rows", priority=True),
        Binding("ctrl+g", "focus_goto", "Go to line", priority=True),
        Binding(":", "focus_command", "Go to line", priority=True),
        Binding("enter", "open_selected", show=False, priority=True),
    ]

    def __init__(
        self,
        root: Path,
        initial_file: Path | None,
        window_size: int,
    ) -> None:
        super().__init__()

        self.root = root.resolve()
        self.current_file = initial_file.resolve() if initial_file else None
        self.window_size = max(window_size, 1)

        self.reader: BaseWindowReader | None = None
        self.window_offset = 0
        self.loaded_rows = 0
        self.left_visible = True
        self._resize_pointer_offset = 0.0

        self.expanded_dirs: set[Path] = {self.root}
        self.visible_file_rows: list[FileRow] = []
        self.project_files: list[Path] = collect_supported_files(self.root)

    def compose(self) -> ComposeResult:
        with Vertical(id="root"):
            with Horizontal(id="main"):
                with Vertical(id="tree_panel"):
                    yield Static(id="tree_header")
                    yield DataTable(id="file_table", zebra_stripes=False)

                yield PaneDivider(id="pane_divider")

                with Vertical(id="data_panel"):
                    yield Static(id="data_header")
                    yield Static(
                        "Open a data file from the file explorer. "
                        "Press ':' or Ctrl+G and enter a line number.",
                        id="status",
                    )
                    yield DataTable(id="data_table", zebra_stripes=True)
                    yield Input(
                        placeholder="Line number, for example 12000",
                        id="command",
                    )

            yield Static(
                "q Quit | r Reload | h Toggle tree | n/PageDown Next chunk | "
                "p/PageUp Previous chunk | g First | Shift+G Last | : Go to line | Enter Open",
                id="help_bar",
            )

    def on_mount(self) -> None:
        command = self.query_one("#command", Input)
        file_table = self.query_one("#file_table", DataTable)
        data_table = self.query_one("#data_table", DataTable)

        command.border_title = "Go to line"

        file_table.cursor_type = "row"
        data_table.cursor_type = "row"

        if hasattr(file_table, "show_header"):
            file_table.show_header = False

        self.render_file_explorer()

        if self.current_file and self.current_file.exists():
            self.load_file(self.current_file)
        else:
            file_table.focus()

    def file_panel_subtitle(self) -> str:
        total = len(self.project_files)

        if total == 0:
            return "0 files"

        if self.current_file and self.current_file.resolve() in self.project_files:
            index = self.project_files.index(self.current_file.resolve()) + 1
            return f"{index} of {total}"

        return f"{total} files"

    def data_panel_subtitle(self) -> str:
        if self.reader is None:
            return "No file loaded"

        total = self.reader.total_rows
        cols = len(self.reader.columns)

        if total is None:
            start = self.window_offset + 1
            end = self.window_offset + self.loaded_rows
            return f"rows {start}-{end} | {cols} cols"

        if total == 0:
            return f"rows 0-0 / 0 | {cols} cols"

        start = self.window_offset + 1
        end = min(self.window_offset + self.loaded_rows, total)
        return f"rows {start}-{end} / {total} | {cols} cols"

    def refresh_panel_metadata(self) -> None:
        tree_header = self.query_one("#tree_header", Static)
        data_header = self.query_one("#data_header", Static)

        tree_header.update(
            Text.assemble(
                ("[2]-Files", "bold"),
                (f"  {self.file_panel_subtitle()}", "dim"),
            )
        )
        data_header.update(
            Text.assemble(
                ("[0]-Data view", "bold"),
                (f"  {self.data_panel_subtitle()}", "dim"),
            )
        )

    def render_file_explorer(self, preserve_path: Path | None = None) -> None:
        file_table = self.query_one("#file_table", DataTable)
        previous_row = self.get_table_cursor_row(file_table)

        if preserve_path is None and 0 <= previous_row < len(self.visible_file_rows):
            preserve_path = self.visible_file_rows[previous_row].path

        self.visible_file_rows = self.build_visible_file_rows()

        file_table.clear(columns=True)
        file_table.add_column("Files")

        for row in self.visible_file_rows:
            file_table.add_row(self.format_file_row(row))

        target_row = 0

        if preserve_path is not None:
            for index, row in enumerate(self.visible_file_rows):
                if row.path == preserve_path:
                    target_row = index
                    break

        self.move_table_cursor(file_table, target_row, 0)
        self.refresh_panel_metadata()

    def build_visible_file_rows(self) -> list[FileRow]:
        rows: list[FileRow] = []

        def visit(path: Path, ancestors_last: tuple[bool, ...]) -> None:
            rows.append(FileRow(path=path, ancestors_last=ancestors_last))

            if not path.is_dir():
                return

            if path not in self.expanded_dirs:
                return

            try:
                children = sorted(
                    path.iterdir(),
                    key=lambda item: (not item.is_dir(), item.name.lower()),
                )
            except PermissionError:
                return

            for index, child in enumerate(children):
                is_last = index == len(children) - 1
                visit(child, ancestors_last + (is_last,))

        visit(self.root, tuple())

        return rows

    def format_file_row(self, row: FileRow) -> Text:
        path = row.path
        prefix = self.format_tree_prefix(row.ancestors_last)
        name = path.name or str(path)

        if path.is_dir():
            expanded = path in self.expanded_dirs
            heavy = path.name in HEAVY_DIRS
            icon = FOLDER_OPEN_ICON if expanded else FOLDER_CLOSED_ICON

            if heavy:
                return Text.assemble(
                    (prefix, "dim"),
                    (f"{icon} ", "magenta"),
                    (f"{name}/", "bold magenta"),
                    ("  heavy", "dim"),
                )

            return Text.assemble(
                (prefix, "dim"),
                (f"{icon} ", "cyan"),
                (f"{name}/", "bold cyan"),
            )

        icon, style = self.file_icon(path)
        is_current = (
            self.current_file is not None and path.resolve() == self.current_file
        )

        if is_current:
            return Text.assemble(
                (prefix, "dim"),
                (f"{icon} ", "bold green"),
                (name, "bold green"),
            )

        return Text.assemble(
            (prefix, "dim"),
            (f"{icon} ", style),
            (name, style),
        )

    def format_tree_prefix(self, ancestors_last: tuple[bool, ...]) -> str:
        if not ancestors_last:
            return ""

        parts: list[str] = []

        for is_last in ancestors_last[:-1]:
            parts.append("   " if is_last else "│  ")

        parts.append("└─ " if ancestors_last[-1] else "├─ ")

        return "".join(parts)

    def file_icon(self, path: Path) -> tuple[str, str]:
        suffix = path.suffix.lower()
        name = path.name.lower()

        if name == ".gitignore":
            return FILE_GIT_ICON, "red"

        if suffix in {".csv", ".data", ".tsv", ".xlsx", ".xls", ".parquet"}:
            return FILE_TABLE_ICON, "green"

        if suffix == ".py":
            return FILE_PYTHON_ICON, "yellow"

        if suffix == ".md":
            return FILE_MARKDOWN_ICON, "blue"

        if suffix in {".toml", ".lock", ".json", ".yaml", ".yml"}:
            return FILE_CONFIG_ICON, "magenta"

        if suffix == ".txt":
            return FILE_DEFAULT_ICON, "white"

        return FILE_DEFAULT_ICON, "white"

    def open_selected_file_row(self) -> None:
        file_table = self.query_one("#file_table", DataTable)
        row_index = self.get_table_cursor_row(file_table)

        if row_index < 0 or row_index >= len(self.visible_file_rows):
            return

        row = self.visible_file_rows[row_index]
        path = row.path

        if path.is_dir():
            if path in self.expanded_dirs:
                if path != self.root:
                    self.expanded_dirs.remove(path)
            else:
                self.expanded_dirs.add(path)

            self.render_file_explorer(preserve_path=path)
            return

        if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            self.set_status(f"Unsupported file type: {path.name}")
            return

        self.load_file(path)

    def on_data_table_row_selected(self, event: Any) -> None:
        table = getattr(event, "data_table", None)

        if table is not None and table.id == "file_table":
            self.open_selected_file_row()

    async def on_key(self, event: Key) -> None:
        focused = self.screen.focused
        file_table = self.query_one("#file_table", DataTable)
        data_table = self.query_one("#data_table", DataTable)

        if isinstance(focused, Input):
            if event.key == "escape":
                focused.value = ""
                data_table.focus()
                event.stop()
            return

        if focused is file_table and event.key == "space":
            self.open_selected_file_row()
            event.stop()
            return

        if event.key in {"n", "pagedown"}:
            self.action_next_window()
            event.stop()
            return

        if event.key in {"p", "pageup"}:
            self.action_previous_window()
            event.stop()
            return

    def load_file(self, path: Path) -> None:
        data_table = self.query_one("#data_table", DataTable)

        try:
            self.reader = create_reader(path, self.window_size)
        except Exception as exc:
            data_table.clear(columns=True)
            self.set_status(f"Could not open file:\n{path}\n{exc}")
            return

        self.current_file = path.resolve()
        self.window_offset = 0

        self.render_file_explorer(preserve_path=self.current_file)
        self.render_data_window()

    def render_data_window(self) -> None:
        data_table = self.query_one("#data_table", DataTable)
        data_table.clear(columns=True)

        if self.reader is None:
            self.refresh_panel_metadata()
            return

        rows = self.reader.read_window(self.window_offset, self.window_size)
        self.loaded_rows = len(rows)

        data_table.add_column("#")

        for column in self.reader.columns:
            data_table.add_column(str(column))

        for index, row in enumerate(rows):
            absolute_row = self.window_offset + index + 1
            normalized_row = normalize_row_width(row, len(self.reader.columns))

            data_table.add_row(
                Text(str(absolute_row), style="dim", justify="right"),
                *[format_cell(value) for value in normalized_row],
            )

        if self.loaded_rows > 0:
            self.move_table_cursor(data_table, 0, 0)

        self.update_status()
        self.refresh_panel_metadata()
        data_table.focus()

    def update_status(self) -> None:
        if self.reader is None or self.current_file is None:
            return

        total = self.reader.total_rows

        if total is None:
            range_text = (
                f"{self.window_offset + 1}-{self.window_offset + self.loaded_rows}"
            )
            total_text = "unknown"
        elif total == 0:
            range_text = "0-0"
            total_text = "0"
        else:
            start = self.window_offset + 1
            end = min(self.window_offset + self.loaded_rows, total)
            range_text = f"{start:,}-{end:,}"
            total_text = f"{total:,}"

        self.set_status(
            f"File: {self.current_file}\n"
            f"Rows shown: {range_text} / {total_text} | "
            f"Window size: {self.window_size:,} | Columns: {len(self.reader.columns):,}\n"
            f"Keys: n/PageDown next · p/PageUp previous · : go to line · "
            f"Enter opens folder/file"
        )

    def set_status(self, message: str) -> None:
        self.query_one("#status", Static).update(message)

    def go_to_line(self, line_number: int) -> None:
        if self.reader is None:
            self.set_status("No file loaded. Open a data file first.")
            return

        if line_number < 1:
            line_number = 1

        if self.reader.total_rows is not None:
            line_number = min(line_number, self.reader.total_rows)

        self.window_offset = max(line_number - 1, 0)
        self.render_data_window()

    def action_focus_command(self) -> None:
        command = self.query_one("#command", Input)
        command.value = ""
        command.focus()

    def action_focus_goto(self) -> None:
        command = self.query_one("#command", Input)
        command.value = ""
        command.focus()

    def action_open_selected(self) -> None:
        self.open_selected_file_row()

    def check_action(self, action: str, parameters: tuple[object, ...]) -> bool | None:
        if action == "open_selected":
            try:
                return self.screen.focused is self.query_one("#file_table", DataTable)
            except Exception:
                return False

        return super().check_action(action, parameters)

    def on_input_submitted(self, event: Input.Submitted) -> None:
        command = event.value.strip()
        event.input.value = ""

        if command == "":
            self.query_one("#data_table", DataTable).focus()
            return

        self.execute_command(command)

    def execute_command(self, command: str) -> None:
        line_number = command.strip().removeprefix(":").strip()

        if line_number.isdigit() and int(line_number) > 0:
            self.go_to_line(int(line_number))
            return

        self.set_status(
            f"Invalid line number: {command}\n"
            "Press ':' and enter a positive whole number, for example 12000."
        )

    def action_reload(self) -> None:
        if self.current_file:
            current_offset = self.window_offset
            self.load_file(self.current_file)
            self.window_offset = current_offset
            self.render_data_window()

    def action_toggle_left(self) -> None:
        panel = self.query_one("#tree_panel", Vertical)
        divider = self.query_one("#pane_divider", PaneDivider)
        self.left_visible = not self.left_visible
        panel.display = self.left_visible
        divider.display = self.left_visible

    def begin_tree_resize(self, screen_x: float) -> None:
        tree_panel = self.query_one("#tree_panel", Vertical)
        self._resize_pointer_offset = screen_x - tree_panel.region.right

    def resize_tree_from_pointer(self, screen_x: float) -> None:
        main = self.query_one("#main", Horizontal)
        tree_panel = self.query_one("#tree_panel", Vertical)
        divider = self.query_one("#pane_divider", PaneDivider)

        available_width = main.size.width
        divider_width = divider.size.width
        max_width = max(
            MIN_TREE_PANEL_WIDTH,
            available_width - divider_width - MIN_DATA_PANEL_WIDTH,
        )
        requested_width = round(screen_x - main.region.x - self._resize_pointer_offset)
        tree_panel.styles.width = max(
            MIN_TREE_PANEL_WIDTH,
            min(requested_width, max_width),
        )

    def action_next_window(self) -> None:
        if self.reader is None:
            return

        total = self.reader.total_rows

        if total is None:
            self.window_offset += self.window_size
        else:
            last_offset = max(total - self.window_size, 0)
            self.window_offset = min(self.window_offset + self.window_size, last_offset)

        self.render_data_window()

    def action_previous_window(self) -> None:
        if self.reader is None:
            return

        self.window_offset = max(self.window_offset - self.window_size, 0)
        self.render_data_window()

    def action_first_window(self) -> None:
        if self.reader is None:
            return

        self.window_offset = 0
        self.render_data_window()

    def action_last_window(self) -> None:
        if self.reader is None or self.reader.total_rows is None:
            return

        self.window_offset = max(self.reader.total_rows - self.window_size, 0)
        self.render_data_window()

    def get_table_cursor_row(self, table: DataTable) -> int:
        coordinate = getattr(table, "cursor_coordinate", None)

        if coordinate is None:
            return 0

        row = getattr(coordinate, "row", None)

        if row is not None:
            return int(row)

        if isinstance(coordinate, tuple) and coordinate:
            return int(coordinate[0])

        return 0

    def move_table_cursor(self, table: DataTable, row: int, column: int) -> None:
        try:
            table.move_cursor(row=row, column=column, animate=False)
        except TypeError:
            table.move_cursor(row=row, column=column)
        except Exception:
            try:
                table.cursor_coordinate = (row, column)
            except Exception:
                pass


def normalize_row_width(row: list[Any], width: int) -> list[Any]:
    if len(row) < width:
        return row + [""] * (width - len(row))

    if len(row) > width:
        return row[: width - 1] + [",".join(map(str, row[width - 1 :]))]

    return row


def format_cell(value: Any) -> Text:
    if value is None:
        return Text("∅", style="dim")

    text = str(value).strip()

    if text == "" or text.lower() in {"nan", "none", "nat"}:
        return Text("∅", style="dim")

    number = try_parse_number(text)

    if number is not None:
        if number.is_integer():
            return Text(str(int(number)), justify="right")

        return Text(f"{number:.6g}", justify="right")

    if len(text) > 48:
        text = text[:47] + "…"

    return Text(text)


def try_parse_number(text: str) -> float | None:
    try:
        return float(text)
    except ValueError:
        return None


def default_initial_file() -> Path | None:
    candidates = [
        Path("data/processed/tennis_matches_enriched.data"),
        Path("data/interim/tennis_matches_cleaned.data"),
        Path("data/interim/tennis_matches_normalized.data"),
        Path("data/interim/tennis_matches_raw.data"),
    ]

    for candidate in candidates:
        if candidate.exists():
            return candidate

    return None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "--root",
        type=Path,
        default=Path("."),
        help="Project root shown in the left file explorer.",
    )

    parser.add_argument(
        "--file",
        type=Path,
        default=None,
        help="Initial data file to open.",
    )

    parser.add_argument(
        "--window-size",
        type=int,
        default=500,
        help="Number of rows kept in memory and displayed at once.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_args()

    initial_file = args.file or default_initial_file()

    app = TennisDataTUI(
        root=args.root,
        initial_file=initial_file,
        window_size=args.window_size,
    )

    app.run()


if __name__ == "__main__":
    main()
